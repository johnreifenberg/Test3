from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, numbers, Alignment, PatternFill
import numpy as np

from backend.models.model import FinancialModel


class ExcelExporter:
    def __init__(self, model: FinancialModel, results: dict, sensitivity: Optional[dict] = None):
        self.model = model
        self.results = results
        self.sensitivity = sensitivity
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.currency_fmt = '"$"#,##0.00'
        self.pct_fmt = "0.00%"

    def _style_header_row(self, ws, row: int, max_col: int) -> None:
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill

    def _style_header_col(self, ws, col: int, min_row: int, max_row: int) -> None:
        for row in range(min_row, max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill

    def create_summary_sheet(self, wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Summary"

        ws.cell(row=1, column=1, value="Inflection Model Summary").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Model: {self.model.name}")
        ws.cell(row=3, column=1, value=f"Forecast Months: {self.model.settings.forecast_months}")
        calc_mode = self.results.get("calculation_mode", "NPV")
        ws.cell(row=4, column=1, value=f"Calculation Mode: {calc_mode}")

        row = 6
        ws.cell(row=row, column=1, value="Key Results").font = self.header_font

        mode = self.results.get("mode", "")
        if mode == "deterministic":
            if calc_mode == "IRR":
                row += 1
                irr = self.results.get("irr")
                ws.cell(row=row, column=1, value="IRR")
                if irr is not None:
                    ws.cell(row=row, column=2, value=irr).number_format = self.pct_fmt
                else:
                    ws.cell(row=row, column=2, value=self.results.get("irr_error", "N/A"))
            else:
                row += 1
                ws.cell(row=row, column=1, value="NPV")
                ws.cell(row=row, column=2, value=self.results["npv"]).number_format = self.currency_fmt
                row += 1
                irr = self.results.get("irr")
                ws.cell(row=row, column=1, value="IRR")
                if irr is not None:
                    ws.cell(row=row, column=2, value=irr).number_format = self.pct_fmt
                else:
                    ws.cell(row=row, column=2, value="N/A")
                row += 1
                ws.cell(row=row, column=1, value="Terminal Value")
                ws.cell(row=row, column=2, value=self.results.get("terminal_value", 0)).number_format = self.currency_fmt
                row += 1
                payback = self.results.get("payback_period")
                ws.cell(row=row, column=1, value="Payback Period")
                if payback is not None:
                    ws.cell(row=row, column=2, value=round(payback, 1)).number_format = "0.0"
                    ws.cell(row=row, column=3, value="months")
                else:
                    ws.cell(row=row, column=2, value="Never")
        elif mode == "monte_carlo":
            if calc_mode == "IRR":
                stats = [
                    ("IRR Mean", "irr_mean"),
                    ("IRR Median", "irr_median"),
                    ("IRR Std Dev", "irr_std"),
                    ("IRR P10", "irr_p10"),
                    ("IRR P25", "irr_p25"),
                    ("IRR P75", "irr_p75"),
                    ("IRR P90", "irr_p90"),
                ]
                for label, key in stats:
                    row += 1
                    ws.cell(row=row, column=1, value=label)
                    val = self.results.get(key)
                    if val is not None:
                        ws.cell(row=row, column=2, value=val).number_format = self.pct_fmt
                    else:
                        ws.cell(row=row, column=2, value="N/A")
                row += 1
                ws.cell(row=row, column=1, value="Failed Simulations")
                ws.cell(row=row, column=2, value=self.results.get("irr_failed_count", 0))
            else:
                stats = [
                    ("NPV Mean", "npv_mean"),
                    ("NPV Median", "npv_median"),
                    ("NPV Std Dev", "npv_std"),
                    ("NPV P10", "npv_p10"),
                    ("NPV P25", "npv_p25"),
                    ("NPV P75", "npv_p75"),
                    ("NPV P90", "npv_p90"),
                ]
                for label, key in stats:
                    row += 1
                    ws.cell(row=row, column=1, value=label)
                    ws.cell(row=row, column=2, value=self.results.get(key, 0)).number_format = self.currency_fmt
                # Payback period stats
                payback_mean = self.results.get("payback_mean")
                if payback_mean is not None:
                    for label, key in [("Payback Mean", "payback_mean"), ("Payback Median", "payback_median"),
                                       ("Payback P10", "payback_p10"), ("Payback P90", "payback_p90")]:
                        row += 1
                        ws.cell(row=row, column=1, value=label)
                        val = self.results.get(key)
                        if val is not None:
                            ws.cell(row=row, column=2, value=round(val, 1)).number_format = "0.0"
                            ws.cell(row=row, column=3, value="months")
                        else:
                            ws.cell(row=row, column=2, value="N/A")
                never_count = self.results.get("payback_never_count", 0)
                if never_count > 0:
                    row += 1
                    ws.cell(row=row, column=1, value="Payback Never Count")
                    ws.cell(row=row, column=2, value=never_count)

        # Assumptions
        row += 2
        ws.cell(row=row, column=1, value="Key Assumptions").font = self.header_font
        if calc_mode == "NPV":
            row += 1
            ws.cell(row=row, column=1, value="Discount Rate Distribution")
            ws.cell(row=row, column=2, value=str(self.model.settings.discount_rate.to_dict()))
            row += 1
            ws.cell(row=row, column=1, value="Terminal Growth Rate")
            ws.cell(row=row, column=2, value=self.model.settings.terminal_growth_rate).number_format = self.pct_fmt
        if self.model.settings.escalation_rate:
            row += 1
            ws.cell(row=row, column=1, value="Escalation Rate Distribution")
            ws.cell(row=row, column=2, value=str(self.model.settings.escalation_rate.to_dict()))

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def create_cashflows_sheet(self, wb: Workbook) -> None:
        """Transposed layout: streams as rows, months as columns."""
        ws = wb.create_sheet("Monthly Cashflows")
        n_months = self.model.settings.forecast_months

        # Header row: "Stream", "Month 0", "Month 1", ...
        ws.cell(row=1, column=1, value="Stream")
        for m in range(n_months):
            ws.cell(row=1, column=m + 2, value=f"Month {m}")
        self._style_header_row(ws, 1, n_months + 1)

        stream_details = self.results.get("stream_details", {})
        total_cashflows = self.results.get("cashflows", [])
        discount_rate = self.results.get("discount_rate", 0.10)
        monthly_rate = discount_rate / 12

        # One row per stream
        current_row = 2
        stream_ids = list(self.model.streams.keys())
        for sid in stream_ids:
            stream_name = self.model.streams[sid].name
            ws.cell(row=current_row, column=1, value=stream_name)
            ws.cell(row=current_row, column=1).font = self.header_font
            cfs = stream_details.get(sid, [0] * n_months)
            for m in range(n_months):
                val = cfs[m] if m < len(cfs) else 0
                ws.cell(row=current_row, column=m + 2, value=val).number_format = self.currency_fmt
            current_row += 1

        # Total row
        ws.cell(row=current_row, column=1, value="Total")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
        for m in range(n_months):
            total = total_cashflows[m] if m < len(total_cashflows) else 0
            ws.cell(row=current_row, column=m + 2, value=total).number_format = self.currency_fmt
        current_row += 1

        # Discounted CF row
        ws.cell(row=current_row, column=1, value="Discounted CF")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
        for m in range(n_months):
            total = total_cashflows[m] if m < len(total_cashflows) else 0
            discount_factor = 1.0 / (1 + monthly_rate) ** m
            ws.cell(row=current_row, column=m + 2, value=total * discount_factor).number_format = self.currency_fmt
        current_row += 1

        # Cumulative NPV row
        ws.cell(row=current_row, column=1, value="Cumulative NPV")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
        cumulative = 0.0
        for m in range(n_months):
            total = total_cashflows[m] if m < len(total_cashflows) else 0
            discount_factor = 1.0 / (1 + monthly_rate) ** m
            cumulative += total * discount_factor
            ws.cell(row=current_row, column=m + 2, value=cumulative).number_format = self.currency_fmt

        # Style the first column and freeze it
        ws.column_dimensions["A"].width = 20
        ws.freeze_panes = "B2"

    def create_streams_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Stream Details")

        headers = [
            "Stream ID", "Name", "Type", "Start Month", "End Month",
            "Amount Distribution", "Adoption Curve",
            "Parent Stream", "Conversion Rate", "Trigger Delay (months)",
            "Periodicity (months)", "Amount Is Ratio",
            "Unit Value", "Market Units",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header_row(ws, 1, len(headers))

        for row_idx, stream in enumerate(self.model.streams.values(), 2):
            ws.cell(row=row_idx, column=1, value=stream.id)
            ws.cell(row=row_idx, column=2, value=stream.name)
            ws.cell(row=row_idx, column=3, value=stream.stream_type.value)
            ws.cell(row=row_idx, column=4, value=stream.start_month)
            ws.cell(row=row_idx, column=5, value=stream.end_month if stream.end_month else "Perpetual")
            ws.cell(row=row_idx, column=6, value=str(stream.amount.to_dict()))
            ws.cell(row=row_idx, column=7, value=str(stream.adoption_curve.to_dict()) if stream.adoption_curve else "None")
            ws.cell(row=row_idx, column=8, value=stream.parent_stream_id or "None")
            ws.cell(row=row_idx, column=9, value=stream.conversion_rate)
            ws.cell(row=row_idx, column=10, value=stream.trigger_delay_months)
            ws.cell(row=row_idx, column=11, value=stream.periodicity_months if stream.periodicity_months else "N/A")
            ws.cell(row=row_idx, column=12, value="Yes" if stream.amount_is_ratio else "No")
            ws.cell(row=row_idx, column=13, value=str(stream.unit_value.to_dict()) if stream.unit_value else "None")
            ws.cell(row=row_idx, column=14, value=str(stream.market_units.to_dict()) if stream.market_units else "None")

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)].width = 20

    def create_sensitivity_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Sensitivity Analysis")

        if not self.sensitivity or not self.sensitivity.get("parameters"):
            ws.cell(row=1, column=1, value="No sensitivity analysis results available")
            return

        headers = ["Parameter", "Stream", "NPV Swing", "NPV Low", "NPV High", "P10 Value", "P90 Value"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header_row(ws, 1, len(headers))

        ws.cell(row=2, column=1, value="Baseline NPV")
        ws.cell(row=2, column=2, value=self.sensitivity["baseline_npv"]).number_format = self.currency_fmt

        for idx, param in enumerate(self.sensitivity["parameters"], 3):
            ws.cell(row=idx, column=1, value=param["parameter_name"])
            ws.cell(row=idx, column=2, value=param["stream_name"])
            ws.cell(row=idx, column=3, value=param["swing"]).number_format = self.currency_fmt
            ws.cell(row=idx, column=4, value=param["npv_low"]).number_format = self.currency_fmt
            ws.cell(row=idx, column=5, value=param["npv_high"]).number_format = self.currency_fmt
            ws.cell(row=idx, column=6, value=param.get("p10_value", ""))
            ws.cell(row=idx, column=7, value=param.get("p90_value", ""))

        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 18

    def create_distribution_sheet(self, wb: Workbook) -> None:
        calc_mode = self.results.get("calculation_mode", "NPV")
        is_irr = calc_mode == "IRR"
        sheet_title = "IRR Distribution" if is_irr else "NPV Distribution"
        dist_key = "irr_distribution" if is_irr else "npv_distribution"
        value_label = "IRR Value" if is_irr else "NPV Value"
        value_fmt = self.pct_fmt if is_irr else self.currency_fmt

        ws = wb.create_sheet(sheet_title)

        if self.results.get("mode") != "monte_carlo":
            ws.cell(row=1, column=1, value=f"Run Monte Carlo simulation to see {sheet_title.lower()}")
            return

        if not self.results.get(dist_key):
            ws.cell(row=1, column=1, value=f"No {sheet_title.lower()} data available")
            return

        # Percentile table
        ws.cell(row=1, column=1, value=sheet_title).font = Font(bold=True, size=14)

        headers = ["Percentile", value_label]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        self._style_header_row(ws, 3, 2)

        dist_data = np.array(self.results[dist_key])
        percentiles = [1, 5, 10, 25, 50, 75, 90, 95, 99]
        for idx, p in enumerate(percentiles, 4):
            ws.cell(row=idx, column=1, value=f"P{p}").number_format = "0%"
            ws.cell(row=idx, column=2, value=float(np.percentile(dist_data, p))).number_format = value_fmt

        # Histogram data
        hist_start = 4 + len(percentiles) + 1
        ws.cell(row=hist_start, column=1, value="Histogram Data").font = Font(bold=True, size=12)

        hist_headers = ["Bin Start", "Bin End", "Frequency"]
        for col, h in enumerate(hist_headers, 1):
            ws.cell(row=hist_start + 1, column=col, value=h)
        self._style_header_row(ws, hist_start + 1, 3)

        # IQR-based trimmed range (wide Tukey fences: 3×IQR)
        q1 = float(np.percentile(dist_data, 25))
        q3 = float(np.percentile(dist_data, 75))
        iqr = q3 - q1
        fence_low = max(float(dist_data.min()), q1 - 3 * iqr)
        fence_high = min(float(dist_data.max()), q3 + 3 * iqr)

        # Adaptive bin count: Rice rule capped at 50
        n_bins = min(50, max(5, int(np.ceil(2 * len(dist_data) ** (1 / 3)))))

        trimmed = dist_data[(dist_data >= fence_low) & (dist_data <= fence_high)]
        counts, bin_edges = np.histogram(trimmed, bins=n_bins, range=(fence_low, fence_high))

        # Count outliers folded into edge bins
        low_outliers = int(np.sum(dist_data < fence_low))
        high_outliers = int(np.sum(dist_data > fence_high))
        counts[0] += low_outliers
        counts[-1] += high_outliers

        for i, count in enumerate(counts):
            row = hist_start + 2 + i
            ws.cell(row=row, column=1, value=float(bin_edges[i])).number_format = value_fmt
            ws.cell(row=row, column=2, value=float(bin_edges[i + 1])).number_format = value_fmt
            ws.cell(row=row, column=3, value=int(count))

        # Note about outliers if any were folded in
        if low_outliers > 0 or high_outliers > 0:
            note_row = hist_start + 2 + len(counts) + 1
            parts = []
            if low_outliers > 0:
                parts.append(f"{low_outliers} below fence")
            if high_outliers > 0:
                parts.append(f"{high_outliers} above fence")
            ws.cell(row=note_row, column=1, value=f"Outliers folded into edge bins: {', '.join(parts)}")

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 18

    def export(self, filepath: str) -> None:
        wb = Workbook()
        self.create_summary_sheet(wb)
        self.create_cashflows_sheet(wb)
        self.create_streams_sheet(wb)
        self.create_sensitivity_sheet(wb)
        self.create_distribution_sheet(wb)
        wb.save(filepath)
